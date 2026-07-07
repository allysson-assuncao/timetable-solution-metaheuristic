import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from src.core.state import TimetableState
from src.utils.color_hash import get_hex_and_font_color

# Style constants
INDISP_BG       = "4A4A4A"
INDISP_FONT     = "FFFFFF"
HEADER_BG       = "1E1E2E"   # Deep navy for row/col headers
HEADER_FONT     = "CDD6F4"   # Soft white
FREE_BG         = "F8F8F8"   # Near-white for free periods
FREE_FONT       = "AAAAAA"   # Muted gray text

THICK_BORDER    = Side(style="medium",  color="2A2A3E")
THIN_BORDER     = Side(style="thin",    color="BBBBBB")
DIAG_BORDER     = Side(style="hair",    color="888888")
NO_BORDER       = Side(style=None)

def auto_width(ws, col_letter: str, min_w: float = 8.0, max_w: float = 22.0):
    """Set column width based on longest cell content, capped between min and max."""
    col_cells = ws[col_letter]
    max_len = max(
        (len(str(c.value).replace("\n", " ")) for c in col_cells if c.value),
        default=8
    )
    ws.column_dimensions[col_letter].width = min(max(max_len * 0.85, min_w), max_w)

def get_border(col_data_idx: int, periodos_por_dia: int) -> Border:
    """col_data_idx is 0-based index within the data columns (excludes col A)."""
    is_day_start = (col_data_idx % periodos_por_dia) == 0
    left_side  = THICK_BORDER if is_day_start else THIN_BORDER
    return Border(left=left_side, right=NO_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)

class ExportManager:
    def __init__(self, state: TimetableState):
        self.state = state
        self.int_to_class_disc = state.int_to_class_disc

    def export_to_excel(self, filepath: str = "Horario_Escolar_Otimizado.xlsx") -> str:
        matrix = self.state.matrix
        M, P = matrix.shape

        # Recuperar IDs legíveis (Professores)
        prof_ids = [p.id_professor for p in self.state.stp_state.professores]
        prof_names = {p.id_professor: p.nome for p in self.state.stp_state.professores}

        # O state garante que a linha 'i' é exatamente o professor no índice 'i' da lista
        row_labels = [prof_names.get(p_id, str(p_id)) for p_id in prof_ids]

        # Montar rótulos das colunas
        dias_letivos = self.state.stp_state.parametros_execucao.dias_letivos
        periodos_por_dia = self.state.stp_state.parametros_execucao.periodos_por_dia
        dias_semana = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]

        col_labels = []
        for d in range(dias_letivos):
            dia = dias_semana[d % 7]
            for p in range(periodos_por_dia):
                col_labels.append(f"{dia} - P{p+1}")
        col_labels = col_labels[:P]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Horário Completo"
        ws.sheet_properties.tabColor = "89B4FA"

        ws.freeze_panes = "B2"

        header_font = Font(name="Calibri", bold=True, size=11, color=HEADER_FONT)
        header_fill = PatternFill("solid", fgColor=HEADER_BG)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        name_font = Font(name="Calibri", bold=True, size=10, color=HEADER_FONT)
        name_fill = PatternFill("solid", fgColor="2A2A3E")
        name_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

        # Header Row
        ws.cell(row=1, column=1, value="Professor").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = header_align

        for j, col_label in enumerate(col_labels):
            cell = ws.cell(row=1, column=j+2, value=col_label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        ws.row_dimensions[1].height = 36
        ws.column_dimensions["A"].width = 28

        indisp_border = Border(
            left=THIN_BORDER, right=THIN_BORDER,
            top=THIN_BORDER,  bottom=THIN_BORDER,
            diagonalDown=True, diagonalUp=True,
            diagonal=DIAG_BORDER
        )

        for i in range(M):
            row_idx = i + 2
            ws.cell(row=row_idx, column=1, value=row_labels[i]).font = name_font
            ws.cell(row=row_idx, column=1).fill = name_fill
            ws.cell(row=row_idx, column=1).alignment = name_align
            ws.row_dimensions[row_idx].height = 42

            for j in range(P):
                col_idx = j + 2
                cell = ws.cell(row=row_idx, column=col_idx)
                val = matrix[i, j]

                if val == -1:
                    cell.fill = PatternFill("solid", fgColor=INDISP_BG)
                    cell.font = Font(name="Calibri", size=9, bold=True, color=INDISP_FONT)
                    cell.value = "✕ Indisponível"
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = indisp_border
                elif val == 0:
                    cell.fill = PatternFill("solid", fgColor=FREE_BG)
                    cell.font = Font(name="Calibri", size=9, color=FREE_FONT)
                    cell.value = "Livre"
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    class_id, disc_id = self.int_to_class_disc.get(val, ("Desconhecido", "Desconhecido"))
                    turma_name = next((t.nome for t in self.state.stp_state.turmas if t.id_turma == class_id), str(class_id))
                    disc_name = next((d.nome for d in self.state.stp_state.disciplinas if d.id_disciplina == disc_id), str(disc_id))
                    
                    seed_str = f"{turma_name}{disc_name}"
                    bg_hex, font_hex = get_hex_and_font_color(seed_str)
                    
                    cell.fill = PatternFill("solid", fgColor=bg_hex.replace("#", ""))
                    cell.font = Font(name="Calibri", size=9, bold=True, color=font_hex.replace("#", ""))
                    cell.value = f"{turma_name}\n({disc_name})"
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Apply borders for day boundaries
        for i in range(M):
            row_idx = i + 2
            for j in range(P):
                col_idx = j + 2
                cell = ws.cell(row=row_idx, column=col_idx)
                val = matrix[i, j]
                
                # Apply borders to cells unless they are indisponivel (they have their own border)
                if val != -1:
                    cell.border = get_border(j, periodos_por_dia)
        
        # apply border to the header row
        for j in range(P):
            col_idx = j + 2
            cell = ws.cell(row=1, column=col_idx)
            cell.border = get_border(j, periodos_por_dia)

        # Auto width calculation
        for j in range(P + 1):
            auto_width(ws, get_column_letter(j + 1))

        wb.save(filepath)
        print(f"Planilha exportada com sucesso em: {filepath}")
        return filepath
